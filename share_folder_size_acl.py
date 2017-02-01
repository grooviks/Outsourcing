# -*- coding: utf-8 -*-
import win32security
import ntsecuritycon as con
import os
import sys
import logging
import shutil
import glob
import wmi
from subprocess import Popen, PIPE
from ldap3 import Server, Connection, ALL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class UserLDAP():
    """Пользователь домена , принимает ФИО пользователя, компания, имя учетной записи, ou """
    def __init__(self, nameuser, company, samaccountname, ou):
        self.__nameuser = nameuser
        self.__company = company if company  is not None else "No company"
        self.__samaccountname = samaccountname
        #получаем иерархию где находится пользователь в структуре AD (OU)
        #выкидываем первый элемент (CN=ФИО) и последние два (имя домена OU=matorin, OU=local)
        self.__ou = ou.split(',')[1:-4:]
        #убираем из каждого элемента списка 'OU=' и оставляем только имя OU
        self.__ou = [x.replace('OU=','') for x in self.__ou ]

    def __str__(self):
        return self.__nameuser
        #, self.__company, self.__samaccountname
    @property
    def nameuser(self):
        return self.__nameuser
    @property
    def company(self):
        return self.__company
    @property
    def samaccountname(self):
        return self.__samaccountname
    @property
    def ou(self):
        return self.__ou

class MyShare():
	''' Общая папка, принимает параметры имя общей папки, путь к ней на сервере, описание, и размер'''
	def __init__(self, name, path, caption, size):
		self.__name = name
		self.__path = path
		self.__caption = caption
		self.__size = size
	@property
	def name(self):
	    return self.__name
	@property
	def path(self):
	    return self.__path
	@property
	def caption(self):
	    return self.__caption
	@property
	def size(self):
	    return self.__size
	
#атрибуты учетных записей
attr_cn = ['cn']
attr_samacct = ['sAMAccountName']
attr_name = ['name']
attr_company = ['company']
attr_ou = ['distinguishedName']

#администратор домегна 
ADMIN = 'Администратор'
#пароль администратора домена 
ADMIN_PASSWORD = '205winPA$$'
#контроллер домена, DNS имя или IP
DC_SERVER = 'matorin-ad'
#путь к выгружаемому xlsx файлу
report_path = 'C:\\scripts\\Report'
#исключенные папки из подсчета
exlcude_share = ['Profiles','Scan','архивы новые', 'it', 'Lotus', 'Подпись', 'files$', 'drv$', 'IMG$', 'print$', 'ОПИРПАК$', 'Consultant' ]


#подключается к AD и выполняет запрос
def QueryLdap(query, attr):

		server = Server(DC_SERVER, get_info=ALL)
		conn = Connection(server, 'cn=' + ADMIN +',cn=users,dc=matorin,dc=local', ADMIN_PASSWORD, auto_bind=True)
		conn.search(search_base ='DC=matorin,DC=local', search_filter = query, attributes = attr,)
		return_list = []
		for entry in conn.response:
			try:
				return_list.append(entry['attributes'])
			except KeyError:
				continue
		return return_list
	

def get_dir_size(start_path):
		total_size = 0
		for dirpath, dirnames, filenames in os.walk(start_path):
			for f in filenames:
				try:
					fp = os.path.join(dirpath, f)
					total_size += os.path.getsize(fp)
				except FileNotFoundError:
					print('FileNotFoundError!!', dirpath)
					continue
				except PermissionError:
					print('PermissionError!!', dirpath)
					continue
		return total_size		

#показывает список прав доступа к файлу\папке, в данном скрипте 
def ShowAce(path,file,globgroups,samacctusers, companyusers, nameusers):
	# List of all file masks that are interesting
	ACCESS_MASKS={ 2032127:"all",
    	1179817:"re",
    	1180086:"f",
    	1179785:"r",
    	1245631:"ch",
    	1073610752:"w"
	}

	#получаем длину дескриптор безопасности (SECURITY_DESCRIPTOR), передаем адрес каталога и
	#список DACL 
	sd = win32security.GetFileSecurity(path, win32security.DACL_SECURITY_INFORMATION)
	dacl = sd.GetSecurityDescriptorDacl() # instead of dacl = win32security.ACL()
	#возвращает количество прав доступа (пользователей или групп которым разрешен\запрещен доступ к файлу или папке)
	ace_count = dacl.GetAceCount()	
	for i in range(0, ace_count):
		rev, access, usersid = dacl.GetAce(i)
		user, group, usertype = win32security.LookupAccountSid('', usersid)
		acc = ACCESS_MASKS.get(access) #права доступа из словаря ACCESS_MASKS
		if user in globgroups:
			print (user)
			userslist = QueryLdap('(&(objectCategory=person)(objectClass=user)(memberOf=CN='+user+',CN=Users,DC=matorin,DC=local))', attr_name+attr_company+attr_samacct)
			for elem in userslist:
				print (elem.get(attr_samacct[0]), elem.get(attr_name[0]))
				file.write('{2} \t {1} \t \n'.format(elem.get(attr_samacct[0]),elem.get(attr_company[0]),elem.get(attr_name[0]),group,acc))
		elif user in samacctusers: 
			print (user)
			i = samacctusers.index(user)
			file.write('\t {2} \t {1} \n'.format(samacctusers[i],companyusers[i],nameusers[i],group,acc))
		else:
			file.write(('\t {1}/{0} \t \n'.format(user,group,acc)))

def FolderACL(path, globgroups, userslist):
	#получаем длину дескриптор безопасности (SECURITY_DESCRIPTOR), передаем адрес каталога и
	#список DACL 
	sd = win32security.GetFileSecurity(path, win32security.DACL_SECURITY_INFORMATION)
	dacl = sd.GetSecurityDescriptorDacl() # instead of dacl = win32security.ACL()
	#возвращает количество прав доступа (пользователей или групп которым разрешен\запрещен доступ к файлу или папке)
	ace_count = dacl.GetAceCount()	
	res = []
	for i in range(0, ace_count):
		#получаем sid пользователя и его права доступа 
		rev, access, usersid = dacl.GetAce(i)
		#получаем имя домена пользователя и имя учетной записи
		user, group, usertype = win32security.LookupAccountSid('', usersid)
		#проверяем пользователя, если это группа доменная то получаем список ее пользователей
		#print(user, group)
		if user in globgroups:
			#запрос в LDAP по имени группы , возврашает список пользователей, получаем 3 атрибута 
			users_in_group = QueryLdap('(&(objectCategory=person)(objectClass=user)(memberOf=CN='+user+',CN=Users,DC=matorin,DC=local))', attr_name+attr_company+attr_samacct)
			res += [j for i in users_in_group for j in userslist if i.get(attr_samacct[0]) == j.samaccountname]
		#если указаны все пользователи, то у всех
		elif group == 'BUILTIN' and user == 'Пользователи':
			print("Все пользователи! %s\%s" % (group, user))
			res.append("%s\%s" % (group, user))
		#если отдельный пользователь то проверяем есть ли он в списке пользователей и добавляем его 
		else: 
			res += [j for j in userslist if j.samaccountname == user] 
	return res

#получаем списки с именами пользователей
def GetListUsers(query):
        listusers = QueryLdap(query,attr_name+attr_company+attr_samacct+attr_ou)
        users = []
        for elem in listusers:
            users.append(UserLDAP(elem.get(attr_name[0]),  elem.get(attr_company[0]), elem.get(attr_samacct[0]), elem.get(attr_ou[0]) ))
        return users

def GetGlobGroups():
	#получаем список групп в домене
	globgroups = []
	listgroups = QueryLdap ('(groupType:1.2.840.113556.1.4.803:=2)', attr_cn)
	for elem in listgroups: 
		globgroups.append(elem.get(attr_cn[0]))
	#print ("GLOB GROUP=",globgroups[0].get('cn'))
	return globgroups

def export_to_txt(uniqcompany, report_path, share_list, share_acl):
		for company in uniqcompany:
			print (company)
			file = open(os.path.join(report_path,company + '.txt'), 'wt')
			for folder in share_list:
				f = 0
				#print(share_acl[folder.caption])
				for user in share_acl[folder.caption]: 
					if user == 'BUILTIN\Пользователи':
						file.write("Все пользователи\n")
						f = 1 
						break
					elif user.company == company: 
						#print (user.nameuser)
						file.write(user.nameuser + "\n")
						f = 1
					
				#ставим проверку что записались все пользователи для нужной компании(если f=1) и можно добавить вывод
				if f: 
					file.write("\nSHARE FOLDER NAME:s {0} Path: {1} Size:{2:.2f} Mb\n ============== \n ".format(folder.caption,folder.path,folder.size))				
			file.close()


def export_to_excel(uniqcompany, report_path, share_list, share_acl):
	wb = Workbook()
	title = ['№','Наименование', 'Доступ', 'Объем папки (Мб)']
	#оптимизировать функцию , писалось на очень скорую руку =( 

	for company in uniqcompany:
		#создаем лист в книге с названием компании 
		#имя листа не может быть больше 31 символа, поэтому обрезаем до 30
		wb.create_sheet(company[0:30])
		print("COMPANY == %s ==" % company)
		ws3 = wb[company[0:30]]
		#задаем ширину столбцов
		ws3.column_dimensions["A"].width = 8
		ws3.column_dimensions["B"].width = 45
		ws3.column_dimensions["C"].width = 38
		ws3.column_dimensions["D"].width = 22

		#первая строка  - шапка, начинаем заполнение со второй
		for col in range(1,len(title)+1):
			ws3.cell(column=col, row=1, value =  title[col-1])

		curr_row = 1
		#счетчик количества папок для компании
		count = 1
		#в списке общих ресурсов выбираем каждую папку, и в словаре share_acl проверяем каждую папку на наличие пользователей нужной компании
		for folder in share_list:
			row_start = curr_row+1
			#ws3['B2'].alignment = Alignment(wrapText=True)
			#обнуляем флаг
			f = 0
			#print(share_acl[folder.caption])
			for user in share_acl[folder.caption]: 
				#если все пользователи имею доступ то вписываем это и выходим из цикла
				if user == 'BUILTIN\Пользователи':
					curr_row += 1
					ws3.cell(column=3, row=curr_row, value="Все пользователи")
					f = 1
					break
				#если пользователь имеет компанию такую же что и выбранная то записываем его ФИО в нужный столбец листа с названием компании 
				elif user.company == company: 
					curr_row += 1
					ws3.cell(column=3, row=curr_row, value=user.nameuser)
					f = 1
			#ставим проверку что записались все пользователи для нужной компании(если f=1) и можно добавить остальные параметры в таблицу
			if f: 
				#номер по счету папки для данной компании
				ws3.cell(column=1, row=curr_row, value=count)
				#объединяем ячейки
				ws3.merge_cells(start_row=row_start,start_column=1,end_row=curr_row,end_column=1)
				#название общей папки
				ws3.cell(column=2, row=curr_row, value=folder.caption)
				ws3.merge_cells(start_row=row_start,start_column=2,end_row=curr_row,end_column=2)
				#размер в МБ
				ws3.cell(column=4, row=curr_row, value="{:.0f}".format(folder.size))
				ws3.merge_cells(start_row=row_start,start_column=4,end_row=curr_row,end_column=4)
				count += 1
	wb.save(os.path.join(report_path,'share_folder.xlsx'))


def main(argv):
		c = wmi.WMI("localhost")
		
		glob_groups = GetGlobGroups()
		#список всех пользователей в группе "Пользователи домена", у которых она задана как основная группа
		users_list = GetListUsers('(&(objectCategory=person)(objectClass=user)(primaryGroupID=513))')
   		#получаем список компаний 
		uniqcompany = {user.company for user in users_list}
		share_list = []
		share_acl = {}
		mb=1048576
		#выбираем общие папки на локальном компьютере кроме тех что есть в списке исключения и системных
		for share in c.Win32_Share():
			if share.type != 0 or share.Name in exlcude_share :
				print (share.Name, share.type, share.Caption)
				continue
			#добавляем в список общих папок каждую папку (класс MyShare)
			share_list.append(MyShare(share.Name, share.Path, share.Caption,get_dir_size(share.Path)/mb))
			#добавляем в словарь, где ключ имя папки - список пользователей имеющих к ней доступ
			share_acl[share.Caption] = FolderACL(share.Path, glob_groups, users_list)

		#export_to_txt(uniqcompany, report_path, share_list, share_acl)

		export_to_excel(uniqcompany, report_path, share_list, share_acl)

		return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))