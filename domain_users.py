# -*- coding: utf-8 -*-

import sys
import os
from ldap3 import Server, Connection, ALL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class UserLDAP():
      """Пользователь домена , принимает ФИО пользователя, компания, имя учетной записи, ou """
    def __init__(self, nameuser, company, samaccountname, ou):
        self.__nameuser = nameuser
        self.__company = company if company  is not None else "No company"
        self.__samaccountname = samaccountname
        #получаем иерархию где находится пользователь в структуре AD (OU)
        #выкидываем первый элемент (CN=ФИО) и последние два (имя домена OU=matorin, OU=local)
        self.__ou = ou.split(',')[1:-4:]
        #убираем из каждого элемента списка 'OU=' и оставляем только имя OU
        self.__ou = [x.replace('OU=','') for x in self.__ou ]

    def __str__(self):
        return self.__nameuser
        #, self.__company, self.__samaccountname
    @property
    def nameuser(self):
        return self.__nameuser
    @property
    def company(self):
        return self.__company
    @property
    def samaccountname(self):
        return self.__samaccountname
    @property
    def ou(self):
        return self.__ou
    

#аттрибуты учетных записей
attr_cn = ['cn']
attr_samacct = ['sAMAccountName']
attr_name = ['name']
attr_company = ['company']
attr_ou = ['distinguishedName']

#администратор домегна 
ADMIN = 'Администратор'
#пароль администратора домена 
ADMIN_PASSWORD = 'пароль администратора'
#контроллер домена, DNS имя или IP
DC_SERVER = 'matorin-ad'
#путь к выгружаемому xlsx файлу
report_path = 'C:\\scripts\\Report'


#подключается к AD и выполняет запрос
def QueryLdap(query, attr):
    server = Server(DC_SERVER, get_info=ALL)
    conn = Connection(server, 'cn=' +  ADMIN + ',cn=users,dc=matorin,dc=local', ADMIN_PASSWORD, auto_bind=True)
    conn.search(search_base ='DC=matorin,DC=local', search_filter = query, attributes = attr,)
    return_list = []
    for entry in conn.response:
        try: 
            return_list.append(entry['attributes'])
        except KeyError:
            continue
    return return_list

#получаем списки с именами пользователей
def GetListUsers(query):
        listusers = QueryLdap(query,attr_name+attr_company+attr_samacct+attr_ou)
        users = []
        for elem in listusers:
            users.append(UserLDAP(elem.get(attr_name[0]),  elem.get(attr_company[0]), elem.get(attr_samacct[0]), elem.get(attr_ou[0]) ))
        return users


def export_to_excel(uniqcompany, users_list_domain, users_list_farm, report_path):
    wb = Workbook()
    #шапка
    title = ['Предоставление доступа к корпоративной файловой системе (доступ в домен), руб', 'Доступ', 'ФИО', 'Отдел']
    #оптимизировать нах!!! 
    for company in uniqcompany:
        #создаем лист в книге с названием компании 
        #имя листа не может быть больше 31 символа, поэтому обрезаем до 30
        wb.create_sheet(company[0:30])
        print("COMPANY == %s ==" % company)
        ws3 = wb[company[0:30]]
        #задаем ширину столбцов
        ws3.column_dimensions["A"].width = 43
        ws3.column_dimensions["B"].width = 40
        ws3.column_dimensions["C"].width = 60
        ws3.column_dimensions["D"].width = 36

        #первая строка  - шапка, начинаем заполнение со второй
        for col in range(1,len(title)+1):
            ws3.cell(column=col, row=1, value =  title[col-1]).alignment = Alignment(wrapText=True)

        curr_row = 1
        start_row = curr_row + 1 
        #счетчик количества папок для компании
        count = 0
        ws3.cell(column=1, row=curr_row+1, value = 'Предоставление доступа и поддержание учетной записи сотрудника в корпоративной файловой системе - доступа в домен (Профиль "Сотрудник"), 490').alignment = Alignment(wrapText=True)
        for user in users_list_domain: 
            #обнуляем флагs
            #если пользователь имеет компанию такую же что и выбранная то записываем его ФИО в нужный столбец листа с названием компании 
            if user.company == company: 
                curr_row += 1
                ws3.cell(column=3, row=curr_row, value=user.nameuser)
                if company == 'МАТОРИН':
                   ws3.cell(column=4, row=curr_row, value=user.ou[0])
                count += 1
        ws3.cell(column=5, row=curr_row, value=count)
        ws3.merge_cells(start_row=start_row,start_column=1,end_row=curr_row,end_column=1)
        ws3.merge_cells(start_row=start_row,start_column=2,end_row=curr_row,end_column=2)
        curr_row += 1
        ws3.cell(column=1, row=curr_row, value= 'Предоставление доступа и поддержание учетной записи сотрудника в корпоративной файловой системе - доступа в домен (Профиль "Уволенный Сотрудник"), 650').alignment = Alignment(wrapText=True)
        curr_row += 1
        ws3.cell(column=1, row=curr_row, value='Поддержание учетной записи сотрудника в корпоративной файловой системе (Профиль "Уволенный Сотрудник"), 200').alignment = Alignment(wrapText=True)
        count = 0
        start_row = curr_row + 1
        ws3.cell(column=1, row=start_row, value = 'Аренда виртуального рабочего места  (Профиль "Ферма"), 1250').alignment = Alignment(wrapText=True)
        for user in users_list_farm:
            if user.company == company:
                curr_row += 1
                ws3.cell(column=3, row=curr_row, value=user.nameuser)
                #если назвавание компании МАТОРИН, то добавляем названивае подразделения
                if company == 'МАТОРИН':
                   ws3.cell(column=4, row=curr_row, value=user.ou[0])
                count += 1
        ws3.cell(column=5, row=curr_row, value=count)
        ws3.merge_cells(start_row=start_row,start_column=1,end_row=curr_row,end_column=1)
        ws3.merge_cells(start_row=start_row,start_column=2,end_row=curr_row,end_column=2)
        curr_row += 1
        ws3.cell(column=1, row=curr_row, value= 'Аренда виртуального рабочего места  (Профиль "Остров"), 7800').alignment = Alignment(wrapText=True)
    wb.save(os.path.join(report_path,'Пользователи домена и фермы.xlsx'))

def main(argv):
    #группа пользователей домена
    group_domain = "Пользователи домена"
    #группы пользователей фермы терминальных серверов
    groups_farm = ["Пользователи RDSFARM", "Пользователи квартплаты", "Only_internet"]
    exclude_ou = ['Управление ИТ']
    #переделать через словарь, ибо сейчас только 2 группы пользователей есть , а возможно и больше будет
    #[x.encode('utf8') for x in groups_farm]
    #список всех пользователей в группе "Пользователи домена", у которых она задана как основная группа
    all_users_list_domain = GetListUsers('(&(objectCategory=person)(objectClass=user)(primaryGroupID=513))')
    all_users_list_farm = []
    #список пользователей фермы, которые включены в группы списка groups_farm
    for group in groups_farm: 
        all_users_list_farm += GetListUsers('(&(objectCategory=person)(objectClass=user)(memberOf=CN='+group+',CN=Users,DC=matorin,DC=local))')

    #получаем уникальные компании
    uniqcompany = {user.company for user in all_users_list_domain}

    #получаем списки без подразделений которые не надо считать в аутсорсинге
    resuslt_users_list_domain = [user for user in all_users_list_domain for ou in exclude_ou if ou not in user.ou ] 
    resuslt_users_list_farm = [user for user in all_users_list_farm for ou in exclude_ou if ou not in user.ou ]

    export_to_excel(uniqcompany, resuslt_users_list_domain, resuslt_users_list_farm, report_path)

    return 0

if __name__ == '__main__':
        sys.exit(main(sys.argv))


