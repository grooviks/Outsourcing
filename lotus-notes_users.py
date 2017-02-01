import win32com.client
import pywintypes
import datetime
import os
import openpyxl
from openpyxl import Workbook
from collections import OrderedDict
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

#путь к отчету
report_path = 'C:\\scripts\\Report'
#адрес сервера LN
notesServer = '172.20.20.2'
#имя базы
notesFile = 'person.nsf'
#пароль пользователя
notesPass = '180505ois'
#файл с пользователями не имеющим доступа
#переделать чтобы выгружались тоже из базы LN
close_access_file = 'close_access.txt'


class Person(object):
	"""attributes - notesname, company, post, fullname, hierarchy, mail, phone"""
	def __init__(self, **kw):
		self.__notesname = kw['notesname']
		self.__company = kw['company'].strip() if kw['company'] != '' else "No company"
		self.__post = kw['post']
		self.__mail = kw['mail']
		self.__fullname = kw['fullname']
		self.__hierarchy = kw['hierarchy']
		self.__phone = kw['phone']
		self.__dismiss = kw['dismiss']
				
	@property
	def notesname(self):
	    return self.__notesname
	@property
	def company(self):
	    return self.__company
	@property
	def post(self):
	    return self.__post
	@property
	def fullname(self):
	    return self.__fullname
	@property
	def hierarchy(self):
	    return self.__hierarchy
	@property
	def phone(self):
	    return self.__phone
	@property
	def mail(self):
	    return self.__mail
	@property
	def dismiss(self):
		return self.__dismiss

def get_view_in_array_views(views_list, view_name ):
	for view in views_list: 
		if view.Name == view_name:
			return view

def document_generator(view):
	document = view.GetFirstDocument()
	while document:
		yield document
		document = view.GetNextDocument(document)

def get_person_list(view):
	#attr_list = ['fldTitle', 'fldPersonId', 'fldOrgName', 'fldHierarchy', 'fldDescription', 'dsp_fldInternalPhone', 'InternetAddress']
	attr_dict = {'notesname' : 'fldPersonId',
	 'post' : 'fldTitle',
	 'company' : 'fldOrgName',
	 'hierarchy' : 'fldHierarchy',
	 'fullname' : 'fldDescription',
	 'phone' : 'dsp_fldInternalPhone',
	 'mail' : 'InternetAddress',
	 'dismiss':'fldDismissdate'
	 }
	person_list = []
	for doc in document_generator(view):
		tmp_dict = {}
		for key, attr in attr_dict.items(): 
			#лотус возврашает нам tuple необходимо взять первый элемент
			tmp_dict[key] = doc.GetItemValue(attr)[0]
		person_list.append(Person(**tmp_dict))	
	return person_list

def dismiss_in_current_month(dismiss_list, cur_date = datetime.date.today()):
	ret_list = []
	print(cur_date)
	for person in dismiss_list: 
		if person.dismiss : 
			if cur_date.year == person.dismiss.year and cur_date.month == person.dismiss.month: 
				print(person.fullname)
				ret_list.append(person)
	return ret_list

def get_company_final_dict(uniq_company, person_list, close_access_list, dismiss_list):
	#определяем списки должностей
	leaderships_list = ["Начальник отдела",  "Главный бухгалтер", "Заместитель руководителя Компании по экономике и финансам"]
	boss_list = ["Руководитель Компании" , "Генеральный директор", "Директор"]
	#исключаем УИТ , так как мы его не считаем в аутсорсинге =) 
	exclude_department = 'ООО "МАТОРИН"\Управление информационных технологий'
	final_dict = {}
	for company in uniq_company: 
		#print("Компания: ", company)
		#company_post_dict = {}
		tmp_boss_list = []
		tmp_leaderships_list = []
		tmp_blocked_list = []
		tmp_person_list = []
		tmp_dismiss_list = []
		company_post_dict = OrderedDict([("Руководители",list()), ("Начальники",list()), ("Сотрудники",list()),
			("Уволенные", list()),
			("Заблокированные", list())])
		for person in person_list: 
			if person.company == company and  exclude_department not in person.hierarchy and person.notesname :
				#print(person.post)
				if person.notesname not in close_access_list : 
					if person.post in boss_list : 
						tmp_boss_list.append(person)
						company_post_dict["Руководители"] = tmp_boss_list
						#print("Начальство: ")
					elif person.post in leaderships_list: 
						tmp_leaderships_list.append(person)
						company_post_dict["Начальники"] = tmp_leaderships_list
					else : 
						tmp_person_list.append(person)
						company_post_dict["Сотрудники"] = tmp_person_list
				else:
					tmp_blocked_list.append(person)
					company_post_dict["Заблокированные"] = tmp_blocked_list
		#company_post_dict["Уволенные"] = [x for x in dismiss_list if x.company == company and x.notesname]
		final_dict[company] = company_post_dict
	return final_dict

def print_all_final_dict(final_dict): 
	for company, posts in final_dict.items(): 
		print("\n\nКомпания: ",company)
		for post, persons in posts.items():
			print (post,":")
			for x in persons:
				print("\t",x.hierarchy, x.fullname)


def export_to_excel(final_dict, uniq_company, report_path):
	#оптимизировать нах!
	wb = Workbook()
	title = ['Профиль, руб.','Доступ ка БД LN и WEB ресурсам','Особенности доступа','ФИО']
	#описание полей для выгрузки в эксель, для каждого профиля - свои описания
	posts_desc = {"Руководители":""" Предоставление и поддержание учетной записи сотрудника в системе 
		Lotus Notes и корпоративным WEB-сервисам (Профиль "Руководитель высшего звена"), 6200""", 
	"Начальники":""" Предоставление и поддержание учетной записи сотрудника в системе Lotus Notes и 
		корпоративным WEB-сервисам (Профиль "Руководитель среднего звена"), 1830 """, 
	"Сотрудники": """ Предоставление и поддержание учетной записи сотрудника в системе Lotus Notes
	 	и корпоративным WEB-сервисам (Профиль "Сотрудник"), 610""",
	 "Уволенные" : """ Предоставление доступа и поддержание учетной записи сотрудника в системе Lotus Notes
		 и корпоративным WEB-сервисам (Профиль "Уволенный сотрудник"), 850""",
	"Заблокированные": """ Поддержание учетной записи сотрудника в системе Lotus Notes 
		и корпоративным WEB-сервисам (Профиль "Заблокированный  сотрудник"), 200"""
	} 

	access_desc_value = '''1. БД LN Заявки \n
						2. БД LN Договоры \r\n
						3. БД LN Сотрудники / Структура организации
						4. БД LN Письма/Поручения
						5. БД LN СМК/Бизнес процессы
						6. БД LN Физ. Лица / юр. Лица
						7. БД LN Объекты/Проекты/оборудование
						8. Личный почтовый ящик
						9.WEB сервис заявки /мобильные заявки
						10. Intranet'''

	access_desc = {"Руководители" : access_desc_value , 
					"Начальники" : access_desc_value, 
					"Сотрудники"  : access_desc_value,
					"Уволенные": "Хранится почтовый фаил сотрудника",
					"Заблокированные" : "Все доступы сотрудника сохранены, но временно заблокированы" }

	feauters_desc = { "Руководители" : ''' Доступ к системам ограничен ИТ ролями:
						1. Автор
						2. Исполнитель
						3.Диспетчер
						4. Куратор
						5. Эксперт
						6. Согласователь
						7. Ген. директор/зам ген. Директора /нач. Управления/нач. блока
						8. Возможно индивидуальные настройки работы системы под конкретного сотрудника''' , 
					"Начальники" : ''' Доступ к системам ограничен ИТ ролями:
						1. Автор
						2. Исполнитель
						3.Диспетчер
						4. Куратор
						5. Эксперт
						6. Согласователь ''', 
					"Сотрудники"  : ''' Можно использовать для сотрудников, которым не нужен доступ к сервисам , 
						но сотрудник должен присутствовать в базах. Например техники , без доступа к ИС''',
					"Уволенные" : "Доступ по запросу(заявке) считается для каждого заявителя отдельно",
					"Заблокированные" : ''' Можно использовать для сотрудников,
					 	которым не нужен доступ к сервисам , но сотрудник должен присутствовать в базах.
					  	Например техники , без доступа к ИС'''
	}

	#for company in uniq_company: 
	for company, posts in final_dict.items():
		#создаем лист в книге с названием компании 
		#имя листа не может быть больше 31 символа, поэтому обрезаем до 30
		wb.create_sheet(company[0:30])
		print("COMPANY == %s ==" % company)
		ws3 = wb[company[0:30]]
		#задаем ширину столбцов
		ws3.column_dimensions["A"].width = 70
		ws3.column_dimensions["B"].width = 42
		ws3.column_dimensions["C"].width = 30
		ws3.column_dimensions["D"].width = 46

		#первая строка  - шапка, начинаем заполнение со второй
		curr_row = 1

		for col in range(1,len(title)+1):
			ws3.cell(column=col, row=curr_row, value =  title[col-1])
		
		
		for key, val in posts.items():
			f=0
			#count = 0
			row_start = curr_row+1
			#устанавливаем значение и добавляем перенос на другую строку
			ws3.cell(column=1, row=row_start, value=posts_desc.get(key)).alignment = Alignment(wrapText=True)
			ws3.cell(column=2, row=row_start, value=access_desc.get(key)).alignment = Alignment(wrapText=True)
			ws3.cell(column=3, row=row_start, value=feauters_desc.get(key)).alignment = Alignment(wrapText=True)

			for lst in val: 
				curr_row += 1
				#count += 1
				ws3.cell(column=4, row=curr_row, value=lst.fullname)	
				#если имеются значения для записи то ставим флаг в 1
				f += 1		
			curr_row += 1 

			if f != 0: 
				print(row_start, curr_row)
				for i in range(1,4): 
					ws3.merge_cells(start_row=row_start,start_column=i,end_row=curr_row,end_column=i)
			ws3.cell(column=5, row=curr_row, value=f)
	wb.save(os.path.join(report_path,'Пользователи LN+Web.xlsx'))


def main():

	#устанавливаем подключение к Lotus сессии
	notesSession = win32com.client.Dispatch("Lotus.NotesSession")
	try:
		#инициализация - ввод пароля
		notesSession.Initialize(notesPass)
		#подключаемся к базе - возвращает объект типа notesDatabase
		db = notesSession.GetDatabase(notesServer,notesFile)
	except pywintypes.com_error:
		raise Exception('Cannot access database using %s on %s' % (notesFile, notesServer))

	#список уволенных сотрудников из представления Сотрудники\Уволенные по фамилиям
	dismiss_list = get_person_list (get_view_in_array_views(db.Views, 'Сотрудники\Уволенные по фамилиям'))
	#список работающих сотрудников из представления Сотрудники\По фамилиям
	person_list = get_person_list (get_view_in_array_views(db.Views, 'Сотрудники\По фамилиям'))
	#считываем файл с заблокированными notes name
	with open(close_access_file, 'r') as f:
		close_access_list = [line.strip() for line in f]
	#сделать через map и lambda
	#находим уволенных в этом месяце - берем текущую дату по дефолту или же передаем в вторым параметром (datetime.datetime(2016,12,30)
	dismiss_list = dismiss_in_current_month(dismiss_list)
	#получаем кортеж уникальных компаний
	uniq_company = {person.company for person in person_list}
	#добавляем уволенных в общий список
	person_list += dismiss_list
	#получаем словарь формата 
	#{'название компании':{'Начальство':[person,person,...], 'Руководители':[person,person,...],'Сотрудники':[person,person,...]}}
	company_final_dict = get_company_final_dict(uniq_company,person_list,close_access_list, dismiss_list)
	#печать на экране итогового словаря debug
	#print_all_final_dict(company_final_dict)
	#экспорт в эксель 
	export_to_excel(company_final_dict, uniq_company, report_path)




if __name__ == "__main__":
    main()