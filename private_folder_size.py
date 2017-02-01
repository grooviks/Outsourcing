# -*- coding: utf-8 -*-

import math
import sys
import os
import glob 
from ldap3 import Server, Connection, ALL
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

class UserLDAP():
    """docstring for UserLDAP"""
    def __init__(self, nameuser, company, samaccountname, ou):
        self.__nameuser = nameuser
        self.__company = company if company  is not None else "No company"
        self.__samaccountname = samaccountname
        #получаем иерархию где находится пользователь в структуре AD (OU)
        #выкидываем первый элемент (CN=ФИО) и последние два (имя домена OU=matorin, OU=local)
        self.__ou = ou.split(',')[1:-4:]
        #убираем из каждого элемента списка 'OU=' и оставляем только имя OU
        self.__ou = [x.replace('OU=','') for x in self.__ou ]

    def __str__(self):
        return self.__nameuser
    @property
    def nameuser(self):
        return self.__nameuser
    @property
    def company(self):
        return self.__company
    @property
    def samaccountname(self):
        return self.__samaccountname
    @property
    def ou(self):
        return self.__ou


#делаем аттрибуты пользователей в домене глобальными, они не меняются и используются в разных функциях
attr_cn = ['cn']
attr_samacct = ['sAMAccountName']
attr_name = ['name']
attr_company = ['company']
attr_ou = ['distinguishedName']

#администратор домегна 
ADMIN = 'Администратор'
#пароль администратора домена 
ADMIN_PASSWORD = '205winPA$$'
#контроллер домена, DNS имя или IP
DC_SERVER = 'matorin-ad'
#путь к выгружаемому xlsx файлу
report_path = 'C:\\scripts\\Report'


#подключается к AD и выполняет запрос
def QueryLdap(query, attr):
    server = Server(DC_SERVER, get_info=ALL)
    conn = Connection(server, 'cn=' +  ADMIN + ',cn=users,dc=matorin,dc=local', ADMIN_PASSWORD, auto_bind=True)
    conn.search(search_base ='DC=matorin,DC=local', search_filter = query, attributes = attr,)
    return_list = []
    for entry in conn.response:
        try: 
            return_list.append(entry['attributes'])
        except KeyError:
            continue
    return return_list

def GetSizeFolder(path):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(path):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                try:
                    total_size += os.path.getsize(fp)
                except (FileNotFoundError, PermissionError):
                    #print("File error!", fp)
                    continue
                
    return total_size

#получаем списки с именами пользователей
def GetListUsers(query):
        listusers = QueryLdap(query,attr_name+attr_company+attr_samacct+attr_ou)
        users = []
        for elem in listusers:
            users.append(UserLDAP(elem.get(attr_name[0]),  elem.get(attr_company[0]), elem.get(attr_samacct[0]), elem.get(attr_ou[0]) ))
        return users

def get_dict_folders(users, server):
    return_dict = {}
    mb=1048576
    for user in users:
        #for folder in os.listdir(path_dir):
        folder = os.path.join(server,user.samaccountname)
        if os.path.exists(folder):
            #получаем размер папки в мегабайтах
            size = GetSizeFolder(folder) // mb
            print(folder)
            print(size)
            return_dict[user]=size
    return return_dict


def export_to_excel(uniqcompany,folders_dict,report_path):
    #оптимизировать нах!
    wb = Workbook()
    title = ['ФИО' , ' Учетная запись',  'Объем личной папки(Мб)' , 'Превышение лимита (Мб)']
    for company in uniqcompany:
        #создаем лист в книге с названием компании 
        #имя листа не может быть больше 31 символа, поэтому обрезаем до 30
        wb.create_sheet(company[0:30])
        print("COMPANY == %s ==" % company)
        ws = wb[company[0:30]]
        #задаем ширину столбцов
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 16

        #первая строка  - шапка, начинаем заполнение со второй
        for col in range(1,len(title)+1):
            ws.cell(column=col, row=1, value =  title[col-1]).alignment = Alignment(wrapText=True)

        curr_row = 1
        excess_size = 0
        for user, size in folders_dict.items(): 
            #если пользователь имеет компанию такую же что и выбранная то записываем его ФИО в нужный столбец листа с названием компании 
            if user.company == company: 
                curr_row += 1
                ws.cell(column=1, row=curr_row, value=user.nameuser)
                ws.cell(column=2, row=curr_row, value=user.samaccountname)
                ws.cell(column=3, row=curr_row, value=size)
                #записываем превышение лимита в 8Гб, если не превышен, то 0
                size = 0 if size < 8192 else size - 8192
                excess_size += size
                ws.cell(column=4, row=curr_row, value=size)
                #если назвавание компании МАТОРИН, то добавляем названивае подразделения
                if company == 'МАТОРИН':
                    ws.cell(column=5, row=curr_row, value=user.ou[0])
        ws.cell(column=4, row=curr_row+1, value=excess_size)
    wb.save(os.path.join(report_path,'private_folder.xlsx'))
       
def main(argv):
    #сервер с личными папками
    server = '//172.20.20.16'
    #исключенные из подсчета OU
    exclude_ou = ['Управление ИТ']     
    all_users_list_domain = GetListUsers('(&(objectCategory=person)(objectClass=user)(primaryGroupID=513))')
    #получаем уникальные компании
    uniqcompany = {user.company for user in all_users_list_domain}
    #получаем списки без подразделений которые не надо считать в аутсорсинге
    resuslt_users_list_domain = [user for user in all_users_list_domain for ou in exclude_ou if ou not in user.ou ] 
    #получаем словарь вида {Пользователь:размер папки} где пользователь UserLDAP class
    folders_dict = get_dict_folders(resuslt_users_list_domain, server)

    export_to_excel (uniqcompany,folders_dict,report_path)
    return 0

if __name__ == '__main__':
        sys.exit(main(sys.argv))


